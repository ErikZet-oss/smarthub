from __future__ import annotations

from pathlib import Path

import pytest

from app.services.inquiry.file_reader import (
    MAX_INQUIRY_ROWS,
    _detect_table_layout,
    read_inquiry_rows_from_bytes,
    read_inquiry_rows_from_path,
)

FIXTURE = Path(__file__).resolve().parent / "fixtures" / "inquiry_specimen.xlsx"


def test_read_csv_single_column() -> None:
    data = b"popis\nskrutka M10x50 DIN933 8.8\nmatice M8\n"
    rows = read_inquiry_rows_from_bytes(data, filename="dopyt.csv")
    assert len(rows) == 2
    assert "M10x50" in rows[0].raw_text


def test_read_csv_with_quantity_column() -> None:
    data = b"ks,text\n10,skrutka M8x40 DIN933\n"
    rows = read_inquiry_rows_from_bytes(data, filename="dopyt.csv")
    assert len(rows) == 1
    assert rows[0].quantity_hint == 10


def test_read_csv_caps_at_max_inquiry_rows() -> None:
    assert MAX_INQUIRY_ROWS == 1500
    body = "\n".join(f"polozka {i:04d}" for i in range(1, MAX_INQUIRY_ROWS + 100))
    rows = read_inquiry_rows_from_bytes(f"popis\n{body}".encode(), filename="dopyt.csv")
    assert len(rows) == MAX_INQUIRY_ROWS


def test_read_csv_named_columns() -> None:
    data = (
        b"Poradove cislo,Nazov polozky,MJ,Pozadovane mnozstvo\n"
        b"1,skrutka M8x40 DIN933,ks,50\n"
        b"2,matica DIN934 M8,ks,100\n"
    )
    rows = read_inquiry_rows_from_bytes(data, filename="dopyt.csv")
    assert len(rows) == 2
    assert "M8x40" in rows[0].raw_text
    assert rows[0].quantity_hint == 50
    assert rows[1].quantity_hint == 100


@pytest.mark.skipif(not FIXTURE.is_file(), reason="inquiry_specimen.xlsx fixture missing")
def test_read_public_procurement_specification_xlsx() -> None:
    rows = read_inquiry_rows_from_path(FIXTURE)
    assert len(rows) >= 100
    assert any("DIN 934" in r.raw_text for r in rows)
    assert any("Závitov" in r.raw_text or "vitov" in r.raw_text.casefold() for r in rows)
    assert rows[0].quantity_hint in (100, 200)
    assert all(r.quantity_hint is not None for r in rows[:20])


@pytest.mark.skipif(not FIXTURE.is_file(), reason="inquiry_specimen.xlsx fixture missing")
def test_detect_layout_for_specification_xlsx() -> None:
    from openpyxl import load_workbook

    wb = load_workbook(FIXTURE, read_only=True, data_only=True)
    ws = wb.active
    matrix = [[("" if c is None else str(c).strip()) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()

    layout = _detect_table_layout(matrix)
    assert layout is not None
    assert layout.text_col == 1
    assert layout.qty_col == 5
    assert layout.data_start_idx == 12
