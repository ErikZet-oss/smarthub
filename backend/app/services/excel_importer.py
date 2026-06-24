from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
import threading
from contextlib import contextmanager
from typing import Any, Callable, Iterator

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.entities import (
    Competitor,
    CompetitorProductMapping,
    FieldMapping,
    Product,
    ProductListItem,
    ProductMapping,
    Supplier,
)

FIELD_DEFAULTS: dict[str, str] = {
    "code": "číslo Smart",
    "norma": "Leading standard",
    "surface": "Surface treatments (long)",
    "diameter": "Diameter [M/Tr]",
    "length": "Length [mm]",
    # Excel: stĺpec V → hlavička „Class“, stĺpec Y → „Money názov“
    "v_class": "Class",
    "y_money_name": "Money názov",
    # Používateľ môže mapovať aj priamo písmenom stĺpca (W).
    "image_filename": "W",
}

# Ak sa v Exceli premenuje hlavička, import skúsi ešte tieto názvy (case-insensitive).
FIELD_HEADER_ALIASES: dict[str, list[str]] = {
    "norma": ["Leading standard", "Leading Standard", "STN"],
    "surface": ["Surface treatments (long)", "Surface treatment (long)"],
    "diameter": ["Diameter [M/Tr]", "Diameter"],
    "length": ["Length [mm]", "Length"],
    "v_class": ["Class"],
    "y_money_name": ["Money názov", "Money nazov"],
    "code": ["číslo Smart", "cislo Smart", "cislo smart"],
}

_FIELD_ATTR_MAP: dict[str, str] = {
    "code": "code_column",
    "norma": "norma_column",
    "surface": "surface_column",
    "diameter": "diameter_column",
    "length": "length_column",
    "v_class": "v_class_column",
    "y_money_name": "y_money_name_column",
    "image_filename": "image_filename_column",
}

_PROFILE_UNIQUE_FIELD_KEYS: tuple[str, ...] = (
    "code",
    "norma",
    "surface",
    "diameter",
    "length",
    "v_class",
    "y_money_name",
    "image_filename",
)
# Profil stĺpcov nepotrebuje unikáty z kódov dodávateľov / EAN — len mapovacie polia.
_PROFILE_MAX_UNIQUE_PER_COLUMN = 1_000
# Ukážka hodnôt v UI — nepotrebujeme celý sheet ak už máme dosť unikátov.
_PROFILE_EARLY_EXIT_MIN_ROWS = 5_000

_EXCEL_IO_LOCK = threading.Lock()


@contextmanager
def excel_io_lock() -> Iterator[None]:
    """Jeden Excel scan/import naraz — na Renderi inak OOM alebo timeout."""
    _EXCEL_IO_LOCK.acquire()
    try:
        yield
    finally:
        _EXCEL_IO_LOCK.release()


def _profile_unique_column_indices(headers: list[str]) -> set[int]:
    """Stĺpce, pre ktoré sa pri profile zbiera unique_values (filtre + mapovanie)."""
    names_cf: set[str] = set()
    for fk in _PROFILE_UNIQUE_FIELD_KEYS:
        default = FIELD_DEFAULTS.get(fk)
        if default:
            names_cf.add(default.casefold())
        for alt in FIELD_HEADER_ALIASES.get(fk, []):
            names_cf.add(alt.casefold())
    letter_headers_cf: set[str] = set()
    for fk in _PROFILE_UNIQUE_FIELD_KEYS:
        default = FIELD_DEFAULTS.get(fk) or ""
        if re.fullmatch(r"[A-Za-z]{1,4}", default):
            letter_headers_cf.add(default.casefold())
    indices: set[int] = set()
    for idx, header in enumerate(headers):
        h = (header or "").strip()
        if not h:
            continue
        h_cf = h.casefold()
        if h_cf in names_cf or h_cf in letter_headers_cf:
            indices.add(idx)
    return indices


_BACKEND_ROOT = Path(__file__).resolve().parent.parent.parent
_REPO_ROOT = _BACKEND_ROOT.parent
_DEFAULT_XLSX = _REPO_ROOT / "data" / "Smart_data_Gamechanger.xlsx"


def resolve_gamechanger_xlsx_path(file_path: str) -> Path:
    """
    Nájde XLSX na disku (absolútna cesta alebo relatívne koreň repa / backend / cwd).
  """
    raw = (file_path or "").strip() or "data/Smart_data_Gamechanger.xlsx"
    p = Path(raw).expanduser()
    candidates: list[Path] = [_DEFAULT_XLSX.resolve()]
    if p.is_absolute():
        candidates.insert(0, p)
    else:
        candidates[:0] = [
            (_REPO_ROOT / p).resolve(),
            (_BACKEND_ROOT / p).resolve(),
            (Path.cwd() / p).resolve(),
        ]
    seen: set[str] = set()
    for c in candidates:
        key = str(c)
        if key in seen:
            continue
        seen.add(key)
        if c.is_file():
            return c
    tried = "\n".join(f"  • {x}" for x in seen)
    raise FileNotFoundError(
        f"Súbor Excel sa nenašiel ({raw!r}). Skúšané cesty:\n{tried}"
    )


def _field_column_name(field_key: str, fm: FieldMapping | None) -> str | None:
    attr = _FIELD_ATTR_MAP.get(field_key)
    if fm and attr:
        raw = getattr(fm, attr, None)
        if raw is not None and str(raw).strip():
            return str(raw).strip()
    return FIELD_DEFAULTS.get(field_key)


@dataclass
class ImportResult:
    products_upserted: int = 0
    products_legacy_removed: int = 0
    suppliers_upserted: int = 0
    mappings_upserted: int = 0
    competitors_upserted: int = 0
    competitor_mappings_upserted: int = 0
    rows_scanned: int = 0
    total_rows: int = 0
    file_resolved: str = ""
    warnings: list[str] = field(default_factory=list)


@dataclass
class ColumnProfileResult:
    sheet: str
    columns: list[str]
    preview_rows: list[dict[str, str]]
    unique_values: dict[str, list[str]]


def _normalize(value: Any) -> str:
    """
    Excel / openpyxl často vracia čísla ako float (strata presnosti pri dlhých kódoch)
    alebo int — pre kódy musíme stabilný text bez „1.23e+15“.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, int):
        return str(value).strip()
    if isinstance(value, float):
        if value != value:  # NaN
            return ""
        rounded = round(value)
        if abs(value - rounded) < 1e-9:
            return str(int(rounded)).strip()
        return str(value).strip()
    return str(value).strip()


def _row_cell(row: Any, idx: int | None) -> str:
    if idx is None or idx < 0:
        return ""
    r = tuple(row) if row is not None else ()
    if idx >= len(r):
        return ""
    return _normalize(r[idx])


def _supplier_index_ci(suppliers: list[Supplier]) -> dict[str, Supplier]:
    out: dict[str, Supplier] = {}
    for s in suppliers:
        key = (s.name or "").strip().casefold()
        if key:
            out[key] = s
    return out


def _competitor_index_ci(competitors: list[Competitor]) -> dict[str, Competitor]:
    out: dict[str, Competitor] = {}
    for c in competitors:
        key = (c.name or "").strip().casefold()
        if key:
            out[key] = c
    return out


def _mapping_entity_name_for_header(
    header: str,
    explicit_header_to_entity: dict[str, str],
    entity_names_longest_first: list[str],
) -> str:
    h = header.strip()
    hit = explicit_header_to_entity.get(h)
    if hit:
        return hit
    low_h = h.casefold()
    for ename in entity_names_longest_first:
        en_cf = ename.casefold()
        if low_h == en_cf:
            return ename
        if ("kód" in low_h or "kod" in low_h) and low_h.startswith(en_cf):
            return ename
    return _to_supplier_name(h)


def _normalized_sheet_headers(first_row: Any) -> list[str]:
    """Rovnaká logika ako pri profile_excel_columns — náhľad a import musia vidieť rovnaké názvy stĺpcov."""
    cells = list(first_row) if first_row is not None else []
    headers = [_normalize(value) for value in cells]
    return [h if h else f"Column {idx + 1}" for idx, h in enumerate(headers)]


def _resolve_header_col_index(headers: list[str], mapped_name: str | None) -> int | None:
    """Nájde stĺpec podľa mapovania — presná zhoda, potom bez rozlišovania veľkosti písmen (ako vo fronte)."""
    if not mapped_name:
        return None
    name = str(mapped_name).strip()
    if not name:
        return None
    for idx, h in enumerate(headers):
        if h == name:
            return idx
    lower = name.lower()
    for idx, h in enumerate(headers):
        if h.lower() == lower:
            return idx
    # Alternatíva: písmeno stĺpca Excelu (A..Z, AA..), napr. "W".
    if re.fullmatch(r"[A-Za-z]{1,4}", name):
        n = 0
        for ch in name.upper():
            n = n * 26 + (ord(ch) - 64)
        idx = n - 1
        if 0 <= idx < len(headers):
            return idx
    return None


def _cislo_smart_col_index(headers: list[str]) -> int | None:
    for alt in FIELD_HEADER_ALIASES.get("code", []) + [FIELD_DEFAULTS["code"]]:
        idx = _resolve_header_col_index(headers, alt)
        if idx is not None:
            return idx
    return None


def _money_short_catalog_col_index(headers: list[str]) -> int | None:
    """Money Katalóg / Money Kód — krátky katalógový kód, nie interné číslo Smart."""
    for idx, h in enumerate(headers):
        low = h.strip().casefold()
        if "money" not in low:
            continue
        if "katal" in low or " kód" in low or low.endswith(" kód") or low.endswith(" kod"):
            return idx
    return None


def _material_number_col_indices(headers: list[str]) -> list[int]:
    """Material Number (without dots) má prednosť pred Material Number."""
    preferred: list[int] = []
    fallback: list[int] = []
    for idx, h in enumerate(headers):
        low = (h or "").strip().casefold()
        if "material number" not in low:
            continue
        if "without dots" in low:
            preferred.append(idx)
        elif low == "material number":
            fallback.append(idx)
    return preferred + fallback


def _resolve_row_internal_code(
    row: Any,
    *,
    primary_idx: int,
    smart_code_idx: int | None,
    money_catalog_idx: int | None,
    material_number_indices: list[int],
) -> str:
    """
    Interný kód produktu — primárne mapovaný stĺpec (číslo Smart).
    Ak chýba, použije Material Number / Money Katalóg (nové riadky v Exceli často nemajú Smart kód).
    """
    primary = _row_cell(row, primary_idx)
    if primary:
        return primary
    if smart_code_idx is not None and smart_code_idx != primary_idx:
        alt = _row_cell(row, smart_code_idx)
        if alt:
            return alt
    for idx in material_number_indices:
        alt = _row_cell(row, idx)
        if alt:
            return alt
    if money_catalog_idx is not None:
        alt = _row_cell(row, money_catalog_idx)
        if alt:
            return alt
    return ""


def _resolve_field_col_index(
    headers: list[str],
    field_key: str,
    fm: FieldMapping | None,
) -> tuple[int | None, str | None, bool]:
    """
    Nájde stĺpec pre mapované pole. Vráti (index, skutočná hlavička, či sa použil alias).
    """
    mapped = _field_column_name(field_key, fm)
    if mapped:
        idx = _resolve_header_col_index(headers, mapped)
        if idx is not None:
            return idx, headers[idx], False
    for alt in FIELD_HEADER_ALIASES.get(field_key, []):
        idx = _resolve_header_col_index(headers, alt)
        if idx is not None:
            return idx, headers[idx], True
    return None, mapped, False


def _sync_field_mapping_columns(
    fm: FieldMapping | None,
    resolved_headers: dict[str, str],
) -> None:
    if not fm:
        return
    for field_key, hdr in resolved_headers.items():
        attr = _FIELD_ATTR_MAP.get(field_key)
        if attr and hdr:
            setattr(fm, attr, hdr)


def _to_supplier_name(header: str) -> str:
    # "Fabory kód" / "Schachermayer nový kód" -> text pred posledným „ kód“
    h = (header or "").strip()
    if not h:
        return ""
    if h.casefold().endswith(" kód"):
        return h[: -len(" kód")].strip()
    return h.rsplit(" ", 1)[0].strip() if " " in h else h


def _mapping_supplier_name_for_header(
    header: str,
    explicit_header_to_supplier: dict[str, str],
    supplier_names_longest_first: list[str],
) -> str:
    """Vyberie názov dodávateľa pre hlavičku stĺpca s kódom (aj po premenovaní v Exceli)."""
    h = header.strip()
    hit = explicit_header_to_supplier.get(h)
    if hit:
        return hit
    low_h = h.casefold()
    if "kód" in low_h or "kod" in low_h:
        for sname in supplier_names_longest_first:
            if low_h.startswith(sname.casefold()):
                return sname
    return _to_supplier_name(h)


def _allocate_supplier_sort_order(session: Session) -> int:
    rows = session.exec(select(Supplier)).all()
    if not rows:
        return 0
    return max((r.sort_order or 0) for r in rows) + 10


def _set_product_field_if_cell_nonempty(
    product: Product,
    field: str,
    row: Any,
    col_idx: int | None,
) -> None:
    """Pri viacerých riadkoch s rovnakým kódom neprepisuj prázdnej bunkou (posledný riadok inak vymaže Class / Money názov)."""
    if col_idx is None:
        return
    r = tuple(row) if row is not None else ()
    if col_idx >= len(r):
        return
    cell = _normalize(r[col_idx])
    if not cell:
        return
    setattr(product, field, cell)


def _excel_image_basename(value: str) -> str:
    raw = (value or "").strip().replace("\\", "/")
    if not raw:
        return ""
    return raw.split("/")[-1].strip()


def import_gamechanger_excel(
    file_path: str,
    session: Session,
    *,
    sheet_name: str = "DIN",
    progress_cb: Callable[[int, int], None] | None = None,
) -> ImportResult:
    with excel_io_lock():
        return _import_gamechanger_excel_locked(
            file_path,
            session,
            sheet_name=sheet_name,
            progress_cb=progress_cb,
        )


def _import_gamechanger_excel_locked(
    file_path: str,
    session: Session,
    *,
    sheet_name: str = "DIN",
    progress_cb: Callable[[int, int], None] | None = None,
) -> ImportResult:
    resolved = resolve_gamechanger_xlsx_path(file_path)
    file_path = str(resolved)
    name = (sheet_name or "DIN").strip() or "DIN"
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return _import_gamechanger_excel_inner(wb, name, file_path, session, progress_cb=progress_cb)
    finally:
        wb.close()


def _import_gamechanger_excel_inner(
    wb,
    name: str,
    file_path: str,
    session: Session,
    *,
    progress_cb: Callable[[int, int], None] | None = None,
) -> ImportResult:
    if name not in wb.sheetnames:
        preview = ", ".join(wb.sheetnames[:30])
        suffix = "…" if len(wb.sheetnames) > 30 else ""
        raise ValueError(
            f"List {name!r} v súbore nie je. Dostupné listy: {preview}{suffix}"
        )

    ws = wb[name]
    total_rows = max(0, int((ws.max_row or 1) - 1))
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = _normalized_sheet_headers(next(rows))

    fm = session.get(FieldMapping, 1)

    field_keys = (
        "code",
        "norma",
        "diameter",
        "length",
        "surface",
        "v_class",
        "y_money_name",
        "image_filename",
    )
    col_by_field: dict[str, int | None] = {}
    resolved_hdr_by_field: dict[str, str] = {}
    mapping_warnings: list[str] = []
    for fk in field_keys:
        idx, hdr, used_alias = _resolve_field_col_index(headers, fk, fm)
        col_by_field[fk] = idx
        if idx is not None and hdr:
            resolved_hdr_by_field[fk] = hdr
            if used_alias and fm:
                old = _field_column_name(fk, fm)
                if old and old != hdr:
                    mapping_warnings.append(
                        f"Mapovanie „{fk}“: v databáze je stĺpec „{old}“, v Exceli sa použije „{hdr}“ "
                        f"(hlavička bola premenovaná). Po importe je mapovanie aktualizované."
                    )

    internal_code_idx = col_by_field["code"]
    norma_idx = col_by_field["norma"]
    diameter_idx = col_by_field["diameter"]
    length_idx = col_by_field["length"]
    surface_idx = col_by_field["surface"]
    v_class_idx = col_by_field["v_class"]
    y_money_name_idx = col_by_field["y_money_name"]
    image_filename_idx = col_by_field["image_filename"]

    if internal_code_idx is None:
        expected = _field_column_name("code", fm)
        raise ValueError(
            f"Stĺpec pre interný kód sa nenašiel (očakávaná hlavička: {expected!r})."
        )

    smart_code_idx = _cislo_smart_col_index(headers)
    money_catalog_idx = _money_short_catalog_col_index(headers)
    if smart_code_idx is not None and internal_code_idx == money_catalog_idx:
        internal_code_idx = smart_code_idx
        col_by_field["code"] = smart_code_idx
        resolved_hdr_by_field["code"] = headers[smart_code_idx]
        mapping_warnings.append(
            f"Mapovanie Kód: v databáze bol krátky stĺpec „{headers[money_catalog_idx]}“, "
            f"import použije „{headers[smart_code_idx]}“ (dlhé interné číslo Smart)."
        )

    leading_idx = _resolve_header_col_index(headers, "Leading standard")
    if leading_idx is not None:
        if norma_idx != leading_idx:
            old_hdr = headers[norma_idx] if norma_idx is not None else "?"
            mapping_warnings.append(
                f"Mapovanie Norma: namiesto „{old_hdr}“ sa pri importe použije "
                f"„{headers[leading_idx]}“ (Leading standard)."
            )
        norma_idx = leading_idx
        col_by_field["norma"] = leading_idx
        resolved_hdr_by_field["norma"] = headers[leading_idx]

    _sync_field_mapping_columns(fm, resolved_hdr_by_field)

    _FIELD_LABELS: dict[str, str] = {
        "norma": "Norma",
        "surface": "Povrch",
        "diameter": "Priemer",
        "length": "Dĺžka",
        "v_class": "Class",
        "y_money_name": "Money názov",
        "image_filename": "Obrázok (W)",
    }
    for fk, label in _FIELD_LABELS.items():
        if col_by_field.get(fk) is None:
            expected = _field_column_name(fk, fm)
            mapping_warnings.append(
                f"Stĺpec „{label}“ sa v Exceli nenašiel "
                f"(mapovanie: {expected!r}). Hodnoty sa pri importe neaktualizujú."
            )

    result = ImportResult()
    result.warnings.extend(mapping_warnings)
    result.file_resolved = file_path
    result.total_rows = total_rows

    code_hdr = headers[internal_code_idx]
    ch_low = code_hdr.casefold()
    if smart_code_idx is not None and internal_code_idx != smart_code_idx:
        result.warnings.append(
            f"Stĺpec „Kód“ je mapovaný na „{code_hdr}“, nie na „{FIELD_DEFAULTS['code']}“. "
            "V tabuľke môžu byť krátke katalógové kódy namiesto dlhého čísla Smart."
        )
    elif money_catalog_idx is not None and internal_code_idx == money_catalog_idx:
        result.warnings.append(
            f"Stĺpec „Kód“ je stále mapovaný na „{code_hdr}“. Nastav v Párovaní Kód → "
            f"„{FIELD_DEFAULTS['code']}“ (nie Money Katalóg / Money Kód)."
        )

    # Stĺpce s kódmi dodávateľov: z DB (code_column) + doplnenie stĺpcov končiacich na " kód".
    supplier_code_columns: list[tuple[int, str]] = []
    seen_indices: set[int] = set()
    # Aktuálna hlavička z Excelu -> UI a ďalší import (po premenovaní stĺpca).
    sync_supplier_code_column: dict[str, str] = {}
    suppliers_rows = session.exec(select(Supplier)).all()
    supplier_names_longest_first = sorted(
        list({(s.name or "").strip() for s in suppliers_rows if (s.name or "").strip()}),
        key=len,
        reverse=True,
    )
    for supplier in suppliers_rows:
        if supplier.code_column:
            col_name = supplier.code_column.strip()
            idx = _resolve_header_col_index(headers, col_name)
            if idx is not None and idx not in seen_indices:
                hdr = headers[idx]
                supplier_code_columns.append((idx, hdr))
                seen_indices.add(idx)
                sname = (supplier.name or "").strip()
                if sname:
                    sync_supplier_code_column[sname] = hdr

    # Starý názov v DB už nesedí, ale hlavička začína na meno dodávateľa a obsahuje kód.
    suppliers_by_len = sorted(
        [s for s in suppliers_rows if (s.name or "").strip()],
        key=lambda s: len((s.name or "").strip()),
        reverse=True,
    )
    for supplier in suppliers_by_len:
        sname = (supplier.name or "").strip()
        if sname in sync_supplier_code_column:
            continue
        low_sn = sname.casefold()
        for idx, header in enumerate(headers):
            if idx in seen_indices:
                continue
            low_h = header.strip().casefold()
            if low_h.startswith(low_sn) and ("kód" in low_h or "kod" in low_h):
                hstrip = header.strip()
                supplier_code_columns.append((idx, hstrip))
                seen_indices.add(idx)
                sync_supplier_code_column[sname] = hstrip
                break

    # Stĺpce s kódmi konkurencie (rovnaký pattern ako dodávatelia).
    competitor_code_columns: list[tuple[int, str]] = []
    sync_competitor_code_column: dict[str, str] = {}
    competitors_rows = session.exec(select(Competitor)).all()
    competitor_names_longest_first = sorted(
        list({(c.name or "").strip() for c in competitors_rows if (c.name or "").strip()}),
        key=len,
        reverse=True,
    )
    for competitor in competitors_rows:
        if competitor.code_column:
            col_name = competitor.code_column.strip()
            idx = _resolve_header_col_index(headers, col_name)
            if idx is not None and idx not in seen_indices:
                hdr = headers[idx]
                competitor_code_columns.append((idx, hdr))
                seen_indices.add(idx)
                cname = (competitor.name or "").strip()
                if cname:
                    sync_competitor_code_column[cname] = hdr
    for idx, header in enumerate(headers):
        if idx in seen_indices:
            continue
        if not header.endswith(" kód"):
            continue
        low_h = header.strip().casefold()
        matched_competitor = False
        for cname in competitor_names_longest_first:
            if low_h.startswith(cname.casefold()):
                competitor_code_columns.append((idx, header))
                seen_indices.add(idx)
                sync_competitor_code_column.setdefault(cname, header)
                matched_competitor = True
                break
        if matched_competitor:
            continue
        supplier_code_columns.append((idx, header))
        seen_indices.add(idx)
        sname = _to_supplier_name(header)
        if sname:
            sync_supplier_code_column.setdefault(sname, header)

    competitors_by_len = sorted(
        [c for c in competitors_rows if (c.name or "").strip()],
        key=lambda c: len((c.name or "").strip()),
        reverse=True,
    )
    for competitor in competitors_by_len:
        cname = (competitor.name or "").strip()
        if cname in sync_competitor_code_column:
            continue
        low_cn = cname.casefold()
        for idx, header in enumerate(headers):
            if idx in seen_indices:
                continue
            low_h = header.strip().casefold()
            hstrip = header.strip()
            if low_h == low_cn or (
                low_h.startswith(low_cn) and ("kód" in low_h or "kod" in low_h)
            ):
                competitor_code_columns.append((idx, hstrip))
                seen_indices.add(idx)
                sync_competitor_code_column[cname] = hstrip
                break

    explicit_header_to_supplier: dict[str, str] = {}
    for supplier in suppliers_rows:
        if supplier.code_column and supplier.code_column.strip():
            explicit_header_to_supplier[supplier.code_column.strip()] = supplier.name

    explicit_header_to_competitor: dict[str, str] = {}
    for competitor in competitors_rows:
        if competitor.code_column and competitor.code_column.strip():
            explicit_header_to_competitor[competitor.code_column.strip()] = competitor.name

    # Akumulácia z Excelu (jeden prechod), potom batch zápis do DB.
    product_payloads: dict[str, dict[str, str | None]] = {}
    mapping_payloads: dict[tuple[str, str], str] = {}
    competitor_mapping_payloads: dict[tuple[str, str], str] = {}
    legacy_short_codes: set[str] = set()
    material_number_indices = _material_number_col_indices(headers)
    skipped_no_code = 0

    for row in rows:
        result.rows_scanned += 1
        if progress_cb is not None and (
            result.rows_scanned <= 25 or result.rows_scanned % 200 == 0
        ):
            progress_cb(result.rows_scanned, result.total_rows)
        internal_code = _resolve_row_internal_code(
            row,
            primary_idx=internal_code_idx,
            smart_code_idx=smart_code_idx,
            money_catalog_idx=money_catalog_idx,
            material_number_indices=material_number_indices,
        )
        if not internal_code:
            skipped_no_code += 1
            continue

        if (
            money_catalog_idx is not None
            and smart_code_idx is not None
            and internal_code_idx == smart_code_idx
        ):
            short_code = _row_cell(row, money_catalog_idx)
            if short_code and short_code != internal_code:
                legacy_short_codes.add(short_code)

        payload = product_payloads.get(internal_code)
        if payload is None:
            payload = {
                "norma": None,
                "diameter": None,
                "length": None,
                "surface": None,
                "v_class": None,
                "y_money_name": None,
                "image_filename": None,
            }
            product_payloads[internal_code] = payload

        r = tuple(row) if row is not None else ()
        idx_to_field = (
            (norma_idx, "norma"),
            (diameter_idx, "diameter"),
            (length_idx, "length"),
            (surface_idx, "surface"),
            (v_class_idx, "v_class"),
            (y_money_name_idx, "y_money_name"),
        )
        for idx, field_name in idx_to_field:
            if idx is None or idx >= len(r):
                continue
            cell = _normalize(r[idx])
            if cell:
                payload[field_name] = cell
        if image_filename_idx is not None:
            if image_filename_idx < len(r):
                img_raw = _normalize(r[image_filename_idx])
                if img_raw:
                    payload["image_filename"] = _excel_image_basename(img_raw) or None
        result.products_upserted += 1

        for col_idx, header in supplier_code_columns:
            supplier_code = _row_cell(row, col_idx)
            if not supplier_code:
                continue

            supplier_name = _mapping_supplier_name_for_header(
                header,
                explicit_header_to_supplier,
                supplier_names_longest_first,
            )
            mapping_payloads[(internal_code, supplier_name)] = supplier_code
            result.mappings_upserted += 1

        for col_idx, header in competitor_code_columns:
            competitor_code = _row_cell(row, col_idx)
            if not competitor_code:
                continue
            competitor_name = _mapping_entity_name_for_header(
                header,
                explicit_header_to_competitor,
                competitor_names_longest_first,
            )
            competitor_mapping_payloads[(internal_code, competitor_name)] = competitor_code
            result.competitor_mappings_upserted += 1

    if skipped_no_code:
        result.warnings.append(
            f"Import: {skipped_no_code} riadkov bez interného kódu "
            "(prázdne číslo Smart, Material Number aj Money Katalóg) sa preskočilo."
        )

    # 1) Produkty: preload + create/update
    existing_products = session.exec(select(Product)).all()
    product_by_code = {p.internal_code: p for p in existing_products}
    for internal_code, payload in product_payloads.items():
        product = product_by_code.get(internal_code)
        if product is None:
            product = Product(internal_code=internal_code)
            session.add(product)
            product_by_code[internal_code] = product
        product.norma = payload["norma"] if isinstance(payload["norma"], str) else product.norma
        product.diameter = (
            payload["diameter"] if isinstance(payload["diameter"], str) else product.diameter
        )
        product.length = payload["length"] if isinstance(payload["length"], str) else product.length
        product.surface = (
            payload["surface"] if isinstance(payload["surface"], str) else product.surface
        )
        product.v_class = payload["v_class"] if isinstance(payload["v_class"], str) else product.v_class
        product.y_money_name = (
            payload["y_money_name"]
            if isinstance(payload["y_money_name"], str)
            else product.y_money_name
        )
        if payload["image_filename"] is not None:
            product.image_filename = (
                payload["image_filename"]
                if isinstance(payload["image_filename"], str)
                else None
            )

    session.flush()

    if legacy_short_codes:
        for short_code in legacy_short_codes:
            if short_code in product_payloads:
                continue
            legacy = product_by_code.get(short_code)
            if legacy is None or legacy.id is None:
                continue
            pid = int(legacy.id)
            for mapping in session.exec(
                select(ProductMapping).where(ProductMapping.product_id == pid)
            ).all():
                session.delete(mapping)
            for list_item in session.exec(
                select(ProductListItem).where(ProductListItem.product_id == pid)
            ).all():
                session.delete(list_item)
            session.delete(legacy)
            product_by_code.pop(short_code, None)
            result.products_legacy_removed += 1
        if result.products_legacy_removed:
            result.warnings.append(
                f"Odstránených {result.products_legacy_removed} starých produktov s krátkym "
                f"katalógovým kódom (Money Katalóg), ktoré mali rovnaký riadok ako číslo Smart."
            )
        session.flush()

    # 2) Dodávatelia: preload + create chýbajúcich (meno bez rozlišovania veľkosti písmen)
    suppliers_rows = session.exec(select(Supplier)).all()
    supplier_by_ci = _supplier_index_ci(suppliers_rows)
    next_sort_order = (
        max((s.sort_order or 0) for s in suppliers_rows) + 10 if suppliers_rows else 0
    )
    for _, supplier_name in mapping_payloads.keys():
        key_ci = (supplier_name or "").strip().casefold()
        if not key_ci or key_ci in supplier_by_ci:
            continue
        canonical = (supplier_name or "").strip()
        supplier = Supplier(
            name=canonical or supplier_name,
            shop_url="",
            username="",
            password="",
            is_connected=False,
            sort_order=next_sort_order,
        )
        next_sort_order += 10
        session.add(supplier)
        supplier_by_ci[key_ci] = supplier
        result.suppliers_upserted += 1

    session.flush()

    suppliers_rows = session.exec(select(Supplier)).all()
    supplier_by_ci = _supplier_index_ci(suppliers_rows)

    # 3) Mappingy: preload a následný upsert bez SELECT v slučke.
    all_mappings = session.exec(select(ProductMapping)).all()
    mapping_by_pair = {(m.product_id, m.supplier_id): m for m in all_mappings}
    skipped_no_supplier = 0
    skipped_no_product = 0
    for (internal_code, supplier_name), supplier_code in mapping_payloads.items():
        product = product_by_code.get(internal_code)
        supplier = supplier_by_ci.get((supplier_name or "").strip().casefold())
        if product is None:
            skipped_no_product += 1
            continue
        if supplier is None:
            skipped_no_supplier += 1
            continue
        if product.id is None or supplier.id is None:
            continue
        key = (int(product.id), int(supplier.id))
        mapping = mapping_by_pair.get(key)
        if mapping is None:
            mapping = ProductMapping(
                product_id=int(product.id),
                supplier_id=int(supplier.id),
                supplier_code=supplier_code,
            )
            session.add(mapping)
            mapping_by_pair[key] = mapping
        else:
            mapping.supplier_code = supplier_code

    if skipped_no_supplier:
        result.warnings.append(
            f"Import: {skipped_no_supplier} väzieb kódu sa neuložilo — v databáze "
            f"nebol dodávateľ podľa mena zo stĺpca (skontroluj názvy dodávateľov vs. hlavičky „… kód“)."
        )
    if skipped_no_product:
        result.warnings.append(
            f"Import: {skipped_no_product} väzieb sa preskočilo — interný kód z Excelu "
            f"nebol v dávke produktov (prázdny alebo nečitateľný stĺpec „Kód“?)."
        )

    # Zosúladenie uloženého názvu stĺpca s hlavičkou v Exceli (viditeľné v apke pri dodávateľovi).
    for supplier in suppliers_rows:
        sname = (supplier.name or "").strip()
        new_hdr = sync_supplier_code_column.get(sname)
        if new_hdr and (supplier.code_column or "").strip() != new_hdr:
            supplier.code_column = new_hdr

    # 4) Konkurencia: create chýbajúcich + mappingy
    competitors_rows = session.exec(select(Competitor)).all()
    competitor_by_ci = _competitor_index_ci(competitors_rows)
    next_comp_sort = (
        max((c.sort_order or 0) for c in competitors_rows) + 10 if competitors_rows else 0
    )
    for _, competitor_name in competitor_mapping_payloads.keys():
        key_ci = (competitor_name or "").strip().casefold()
        if not key_ci or key_ci in competitor_by_ci:
            continue
        canonical = (competitor_name or "").strip()
        competitor = Competitor(
            name=canonical or competitor_name,
            shop_url="",
            sort_order=next_comp_sort,
            is_active=False,
        )
        next_comp_sort += 10
        session.add(competitor)
        competitor_by_ci[key_ci] = competitor
        result.competitors_upserted += 1

    session.flush()
    competitors_rows = session.exec(select(Competitor)).all()
    competitor_by_ci = _competitor_index_ci(competitors_rows)
    all_comp_mappings = session.exec(select(CompetitorProductMapping)).all()
    comp_mapping_by_pair = {
        (m.product_id, m.competitor_id): m for m in all_comp_mappings
    }
    skipped_no_competitor = 0
    for (internal_code, competitor_name), competitor_code in competitor_mapping_payloads.items():
        product = product_by_code.get(internal_code)
        competitor = competitor_by_ci.get((competitor_name or "").strip().casefold())
        if product is None or product.id is None:
            continue
        if competitor is None:
            skipped_no_competitor += 1
            continue
        if competitor.id is None:
            continue
        key = (int(product.id), int(competitor.id))
        mapping = comp_mapping_by_pair.get(key)
        if mapping is None:
            mapping = CompetitorProductMapping(
                product_id=int(product.id),
                competitor_id=int(competitor.id),
                competitor_code=competitor_code,
            )
            session.add(mapping)
            comp_mapping_by_pair[key] = mapping
        else:
            mapping.competitor_code = competitor_code

    if skipped_no_competitor:
        result.warnings.append(
            f"Import: {skipped_no_competitor} väzieb kódu konkurencie sa neuložilo — "
            f"v databáze nebola konkurencia podľa mena zo stĺpca."
        )

    for competitor in competitors_rows:
        cname = (competitor.name or "").strip()
        new_hdr = sync_competitor_code_column.get(cname)
        if new_hdr and (competitor.code_column or "").strip() != new_hdr:
            competitor.code_column = new_hdr

    session.commit()
    if progress_cb is not None:
        progress_cb(result.rows_scanned, result.total_rows)
    return result


def profile_excel_columns(
    file_path: str,
    sheet_name: str = "DIN",
    max_unique_values: int = _PROFILE_MAX_UNIQUE_PER_COLUMN,
    max_scan_rows: int = 500_000,
    preview_row_count: int = 8,
) -> ColumnProfileResult:
    with excel_io_lock():
        return _profile_excel_columns_locked(
            file_path,
            sheet_name=sheet_name,
            max_unique_values=max_unique_values,
            max_scan_rows=max_scan_rows,
            preview_row_count=preview_row_count,
        )


def _profile_excel_columns_locked(
    file_path: str,
    sheet_name: str = "DIN",
    max_unique_values: int = _PROFILE_MAX_UNIQUE_PER_COLUMN,
    max_scan_rows: int = 500_000,
    preview_row_count: int = 8,
) -> ColumnProfileResult:
    file_path = str(resolve_gamechanger_xlsx_path(file_path))
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' was not found in workbook.")

    ws = wb[sheet_name]
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = _normalized_sheet_headers(next(rows))
    unique_column_indices = _profile_unique_column_indices(headers)

    preview_rows: list[dict[str, str]] = []
    unique_sets = [set() for _ in headers]

    scanned = 0
    rows_without_new_unique = 0
    try:
        for row in rows:
            scanned += 1
            cells = tuple(row) if row is not None else ()
            if len(preview_rows) < preview_row_count:
                preview_rows.append(
                    {
                        headers[idx]: _normalize(cells[idx]) if idx < len(cells) else ""
                        for idx in range(len(headers))
                    }
                )

            added_any = False
            for idx in unique_column_indices:
                if len(unique_sets[idx]) >= max_unique_values:
                    continue
                value = _normalize(cells[idx]) if idx < len(cells) else ""
                if not value:
                    continue
                before = len(unique_sets[idx])
                unique_sets[idx].add(value)
                if len(unique_sets[idx]) > before:
                    added_any = True

            if added_any:
                rows_without_new_unique = 0
            else:
                rows_without_new_unique += 1

            if scanned >= _PROFILE_EARLY_EXIT_MIN_ROWS and rows_without_new_unique >= 2_000:
                break
            if all(
                len(unique_sets[idx]) >= max_unique_values for idx in unique_column_indices
            ):
                break
            if max_scan_rows > 0 and scanned >= max_scan_rows:
                break
    finally:
        wb.close()

    unique_values = {
        headers[idx]: sorted(unique_sets[idx])
        for idx in range(len(headers))
        if idx in unique_column_indices and unique_sets[idx]
    }
    return ColumnProfileResult(
        sheet=sheet_name,
        columns=headers,
        preview_rows=preview_rows,
        unique_values=unique_values,
    )
