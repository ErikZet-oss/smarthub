"""Po deployi nového XLSX môže databáza ostať pozadu — doplní import na pozadí."""

from __future__ import annotations

import logging
import os
import threading

from openpyxl import load_workbook
from sqlalchemy import func
from sqlmodel import Session, select

from app.db import engine
from app.models.entities import Product
from app.services.excel_importer import import_gamechanger_excel, resolve_gamechanger_xlsx_path

logger = logging.getLogger(__name__)

_DEFAULT_SHEET = "DIN"
# Ak v DB chýba aspoň toľko riadkov oproti Excelu, spustí sa sync (nové normy / produkty).
_MIN_MISSING_ROWS = 5000
# Po štarte nechaj API najprv obslúžiť UI (Načítať stĺpce), až potom ťažký import.
_STARTUP_SYNC_DELAY_SEC = 180.0


def _excel_din_row_count(path: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if _DEFAULT_SHEET not in wb.sheetnames:
            return 0
        ws = wb[_DEFAULT_SHEET]
        return max(0, int((ws.max_row or 1) - 1))
    finally:
        wb.close()


def _db_product_count() -> int:
    with Session(engine) as session:
        count = session.exec(select(func.count()).select_from(Product)).one()
        return int(count or 0)


def _run_excel_import(path: str) -> None:
    try:
        from app.db import create_db_and_tables, migrate_schema

        create_db_and_tables()
        migrate_schema()
        with Session(engine) as session:
            result = import_gamechanger_excel(
                path,
                session,
                sheet_name=_DEFAULT_SHEET,
            )
            session.commit()
        # Neskorý import — routes je už načítaný pri štarte API.
        from app.api.routes import _filter_opts_cache_invalidate

        _filter_opts_cache_invalidate()
        logger.info(
            "Excel startup sync hotový: %s produktov, %s riadkov (súbor %s)",
            result.products_upserted,
            result.rows_scanned,
            result.file_resolved,
        )
    except Exception:
        logger.exception("Excel startup sync zlyhal")


def _start_excel_sync_if_stale() -> None:
    # Auto-import pri štarte je VYPNUTÝ default-ne. Veľký katalóg (desiatky MB,
    # 64k+ riadkov) by na malom hostingu (Render free, 512 MB) pri parse-ovaní
    # vyčerpal pamäť a zhodil celú službu. Dáta sa do produkcie dostávajú cez
    # manuálny import alebo migráciu DB. Zapneš výslovne premennou prostredia.
    if os.getenv("SMARTHUB_ENABLE_AUTO_EXCEL_SYNC", "").strip().lower() not in (
        "1",
        "true",
        "yes",
    ):
        return
    if os.getenv("SMARTHUB_DISABLE_AUTO_EXCEL_SYNC", "").strip().lower() in (
        "1",
        "true",
        "yes",
    ):
        return
    try:
        path = str(resolve_gamechanger_xlsx_path(""))
    except FileNotFoundError:
        return

    try:
        excel_rows = _excel_din_row_count(path)
        db_count = _db_product_count()
    except Exception:
        logger.exception("Excel startup sync: nepodarilo sa zistiť stav DB/Excel")
        return

    if excel_rows <= 0:
        return
    missing = excel_rows - db_count
    if missing < _MIN_MISSING_ROWS:
        return

    logger.info(
        "Excel startup sync: DB má %s produktov, Excel list %s má ~%s riadkov "
        "(chýba ~%s) — spúšťam import na pozadí.",
        db_count,
        _DEFAULT_SHEET,
        excel_rows,
        missing,
    )
    thread = threading.Thread(
        target=_run_excel_import,
        args=(path,),
        name="excel-startup-sync",
        daemon=True,
    )
    thread.start()


def schedule_excel_sync_if_stale() -> None:
    timer = threading.Timer(_STARTUP_SYNC_DELAY_SEC, _start_excel_sync_if_stale)
    timer.daemon = True
    timer.start()
