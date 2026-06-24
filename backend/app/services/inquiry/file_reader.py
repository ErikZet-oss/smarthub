from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.schemas.inquiry import InquiryInputRow

MAX_INQUIRY_ROWS = 1500
_HEADER_SCAN_ROWS = 60

_TEXT_HEADER_HINTS = (
    "názov položky",
    "nazov polozky",
    "názov materiálu",
    "nazov materialu",
    "názov",
    "nazov",
    "položka",
    "polozka",
    "popis",
    "specifik",
    "materiál",
    "material",
    "spojovac",
    "produkt",
    "item",
    "description",
    "text",
    "dopyt",
    "opis",
)

_SPEC_HEADER_HINTS = (
    "technická špecifikácia",
    "technicka specifikacia",
    "technická specifik",
    "technicka specifik",
    "specifikácia materiálu",
    "specifikacia materialu",
    "specifik",
)

_QTY_HEADER_HINTS = (
    "požadované množstvo",
    "pozadovane mnozstvo",
    "spotreba",
    "historickej spotreby",
    "množstvo",
    "mnozstvo",
    "počet",
    "pocet",
    "quantity",
    "qty",
    "amount",
)

_QTY_HEADER_EXCLUDE = (
    "cena",
    "price",
    "dph",
    "eur",
    "sadzba",
    "bez dph",
    "s dph",
    "merná",
    "merna",
    "jednotka",
)

_FOOTER_TEXT_HINTS = (
    "celková cena",
    "celkova cena",
    "meno, priezvisko",
    "podpis",
    "ponuka",
    "dátum",
    "datum",
)


@dataclass(frozen=True)
class _TableLayout:
    header_row_idx: int
    text_col: int
    qty_col: int | None
    data_start_idx: int
    spec_col: int | None = None


def read_inquiry_rows_from_bytes(
    data: bytes,
    *,
    filename: str,
) -> list[InquiryInputRow]:
    name = (filename or "").strip().lower()
    if name.endswith(".csv"):
        return _read_csv(data)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _read_xlsx(data)
    raise ValueError("Podporované formáty: .xlsx, .xlsm, .csv")


def read_inquiry_rows_from_path(path: str | Path) -> list[InquiryInputRow]:
    p = Path(path)
    return read_inquiry_rows_from_bytes(p.read_bytes(), filename=p.name)


def read_inquiry_rows_from_text(text: str) -> list[InquiryInputRow]:
    """Riadky z textu prilepeného z Excelu (bunky oddelené tabulátorom / „;")."""
    return _rows_from_matrix(_matrix_from_pasted_text(text))


def _matrix_from_pasted_text(text: str) -> list[list[str]]:
    raw = (text or "").replace("\r\n", "\n").replace("\r", "\n").strip("\n")
    if not raw.strip():
        return []
    delimiter = _detect_paste_delimiter(raw)
    if delimiter is None:
        return [[line.strip()] for line in raw.split("\n") if line.strip()]
    # csv.reader zvládne aj bunky s úvodzovkami / viacriadkové bunky z Excelu.
    reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
    return [[_cell_to_str(cell) for cell in row] for row in reader]


def _detect_paste_delimiter(raw: str) -> str | None:
    """Excel kopíruje stĺpce tabulátorom; slovenské CSV často „;". Čiarku ignorujeme
    (býva priamo v popise položky), nech sa text nerozbije na viac stĺpcov."""
    if "\t" in raw:
        return "\t"
    lines = [ln for ln in raw.split("\n") if ln.strip()]
    if lines and sum(1 for ln in lines if ";" in ln) >= max(1, len(lines) // 2):
        return ";"
    return None


def _read_csv(data: bytes) -> list[InquiryInputRow]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return _rows_from_matrix(list(reader))


def _read_xlsx(data: bytes) -> list[InquiryInputRow]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        matrix: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            matrix.append([_cell_to_str(c) for c in row])
    finally:
        wb.close()
    return _rows_from_matrix(matrix)


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _rows_from_matrix(matrix: list[list[str]]) -> list[InquiryInputRow]:
    if not matrix:
        return []

    layout = _detect_table_layout(matrix)
    if layout is not None:
        return _rows_from_layout(matrix, layout)

    col_idx = _pick_text_column(matrix)
    out: list[InquiryInputRow] = []
    start_row = 1 if _looks_like_header(matrix[0]) else 0

    for row in matrix[start_row:]:
        if len(out) >= MAX_INQUIRY_ROWS:
            break
        text = row[col_idx].strip() if col_idx < len(row) else ""
        if not text or _looks_like_footer_text(text):
            continue
        qty = _parse_quantity_from_row(row, skip_col=col_idx)
        out.append(
            InquiryInputRow(
                row_index=len(out) + 1,
                raw_text=text,
                quantity_hint=qty,
            )
        )
    return out


def _rows_from_layout(matrix: list[list[str]], layout: _TableLayout) -> list[InquiryInputRow]:
    out: list[InquiryInputRow] = []
    for row in matrix[layout.data_start_idx :]:
        if len(out) >= MAX_INQUIRY_ROWS:
            break
        text = row[layout.text_col].strip() if layout.text_col < len(row) else ""
        if layout.spec_col is not None and layout.spec_col < len(row):
            spec = row[layout.spec_col].strip()
            if spec and spec.casefold() != text.casefold():
                text = f"{text} — {spec}" if text else spec
        if not text or _looks_like_footer_text(text):
            continue
        if _looks_like_subheader_row(row):
            continue
        qty: int | None = None
        if layout.qty_col is not None and layout.qty_col < len(row):
            qty = _parse_quantity_cell(row[layout.qty_col])
        if qty is None:
            qty = _parse_quantity_from_row(row, skip_col=layout.text_col)
        out.append(
            InquiryInputRow(
                row_index=len(out) + 1,
                raw_text=text,
                quantity_hint=qty,
            )
        )
    return out


def _detect_table_layout(matrix: list[list[str]]) -> _TableLayout | None:
    best_row_idx = -1
    best_text_col = -1
    best_spec_col: int | None = None
    best_qty_col: int | None = None
    best_score = 0

    scan_limit = min(len(matrix), _HEADER_SCAN_ROWS)
    for row_idx in range(scan_limit):
        row = matrix[row_idx]
        text_col, text_score = _best_column_for_hints(row, _TEXT_HEADER_HINTS)
        spec_col, spec_score = _best_column_for_hints(row, _SPEC_HEADER_HINTS)
        qty_col, qty_score = _best_column_for_hints(
            row,
            _QTY_HEADER_HINTS,
            exclude=_QTY_HEADER_EXCLUDE,
        )
        if text_col < 0 or text_score <= 0:
            continue
        score = text_score + (spec_score if spec_col >= 0 else 0)
        score += qty_score * 2 if qty_col >= 0 else 0
        if score > best_score:
            best_score = score
            best_row_idx = row_idx
            best_text_col = text_col
            best_spec_col = spec_col if spec_col >= 0 and spec_score > 0 else None
            best_qty_col = qty_col if qty_col >= 0 and qty_score > 0 else None

    if best_row_idx < 0 or best_text_col < 0 or best_score < 4:
        return None

    data_start = best_row_idx + 1
    while data_start < len(matrix):
        row = matrix[data_start]
        text = row[best_text_col].strip() if best_text_col < len(row) else ""
        if len(text) >= 8 and not _looks_like_footer_text(text):
            break
        data_start += 1

    return _TableLayout(
        header_row_idx=best_row_idx,
        text_col=best_text_col,
        qty_col=best_qty_col,
        data_start_idx=data_start,
        spec_col=best_spec_col,
    )


def _normalize_header(cell: str) -> str:
    return " ".join(cell.casefold().replace("\n", " ").replace("\r", " ").split())


def _header_match_score(
    cell: str,
    hints: tuple[str, ...],
    *,
    exclude: tuple[str, ...] = (),
) -> int:
    norm = _normalize_header(cell)
    if not norm:
        return 0
    for ex in exclude:
        if ex in norm:
            return 0
    score = 0
    for hint in hints:
        if hint in norm:
            score += len(hint)
    return score


def _best_column_for_hints(
    row: list[str],
    hints: tuple[str, ...],
    *,
    exclude: tuple[str, ...] = (),
) -> tuple[int, int]:
    best_idx = -1
    best_score = 0
    for idx, cell in enumerate(row):
        score = _header_match_score(cell, hints, exclude=exclude)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx, best_score


def _looks_like_subheader_row(row: list[str]) -> bool:
    cells = [c.strip() for c in row if c.strip()]
    if not cells:
        return True
    if all(re.fullmatch(r"[A-Z]{1,2}", c, flags=re.IGNORECASE) for c in cells):
        return True
    if all(re.fullmatch(r"[A-Z]{1,2}/100\s*x\s*[A-Z]{1,2}", c, flags=re.IGNORECASE) for c in cells if c):
        return True
    joined = _normalize_header(" ".join(cells))
    if joined in {"a b c d e f g h i j k l m n", "a b e f g h i j l m"}:
        return True
    return False


def _looks_like_footer_text(text: str) -> bool:
    norm = _normalize_header(text)
    if len(norm) < 3:
        return True
    return any(h in norm for h in _FOOTER_TEXT_HINTS)


def _pick_text_column(matrix: list[list[str]]) -> int:
    """Fallback: stĺpec s najdlhšími textami (typicky popis položky)."""
    if not matrix:
        return 0
    max_cols = max(len(r) for r in matrix)
    best_idx = 0
    best_score = -1.0
    for col in range(max_cols):
        texts = [
            r[col].strip()
            for r in matrix
            if col < len(r) and r[col].strip() and not _looks_like_footer_text(r[col])
        ]
        if not texts:
            continue
        avg_len = sum(len(t) for t in texts) / len(texts)
        if avg_len > best_score:
            best_score = avg_len
            best_idx = col
    return best_idx


def _looks_like_header(row: list[str]) -> bool:
    joined = _normalize_header(" ".join(row))
    hints = ("popis", "položka", "polozka", "text", "dopyt", "názov", "nazov", "item")
    return any(h in joined for h in hints)


def _parse_quantity_cell(cell: str) -> int | None:
    s = cell.strip()
    if not s:
        return None
    try:
        if any(ch in s for ch in ".,"):
            val = float(s.replace(",", ".").replace(" ", ""))
            if val.is_integer() and 0 < val < 1_000_000:
                return int(val)
            return None
        if s.isdigit():
            q = int(s)
            if 0 < q < 1_000_000:
                return q
    except (TypeError, ValueError):
        return None
    return None


def _parse_quantity_from_row(row: list[str], *, skip_col: int) -> int | None:
    for idx, cell in enumerate(row):
        if idx == skip_col:
            continue
        q = _parse_quantity_cell(cell)
        if q is not None:
            return q
    return None
